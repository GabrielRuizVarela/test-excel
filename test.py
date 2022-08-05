# Copy column A from Book1.xlsx and column B from Book2.xlsx then created Book3.xlsx 
# concat column A and column B then save to Book3.xlsx in column D

import openpyxl

Book1 = openpyxl.load_workbook('Book1.xlsx')
Book2 = openpyxl.load_workbook('Book2.xlsx')

#create Book3
Book3 = openpyxl.Workbook()

#get column A from Book1.xlsx
sheet1 = Book1.active
colA = sheet1['A']

# get column B from Book2.xlsx
sheet2 = Book2.active
colB = sheet2['B']

# concat column A and column B then save to Book3.xlsx in column D
sheet3 = Book3.active
#concat column A and column B
for i in range(1, len(colA)):
    sheet3.cell(row=i, column=3).value = colA[i].value
    #add column B of Book2.xlsx to column D of Book3.xlsx below existing cells
    sheet3.cell(row=i+len(colA)-1, column=3).value = colB[i].value




#save and close Book3.xlsx  
Book3.save('Book3.xlsx')
Book3.close()
Book1.close()
Book2.close()

print('Done')